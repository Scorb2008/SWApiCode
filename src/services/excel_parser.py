from decimal import Decimal

from openpyxl import load_workbook


def parse_excel(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    accounts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        login = str(row[0]).strip() if row[0] else ""
        password = str(row[1]).strip() if row[1] else ""
        size = str(row[2]).strip() if row[2] else ""
        price = row[3] if len(row) > 3 and row[3] is not None else 0
        status = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        if not login or not password or not size:
            continue

        try:
            price_decimal = Decimal(str(price)).quantize(Decimal("0.01"))
        except (ValueError, TypeError):
            price_decimal = Decimal("0")

        accounts.append({
            "login": login,
            "password": password,
            "size": size,
            "price": price_decimal,
            "status": status,
        })

    return accounts
